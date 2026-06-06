import openpyxl
import json
import re

# Load the Excel file
wb = openpyxl.load_workbook('TrainerMentors_Complete_Course_Catalog.xlsx')
ws = wb.active

# Extract data
courses_list = []
headers = []

# Get headers from first row
for cell in ws[1]:
    headers.append(cell.value)

# Get data from remaining rows
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:  # If first column has data
        course = {}
        for i, header in enumerate(headers):
            if i < len(row):
                course[header] = row[i]
        courses_list.append(course)

print(f"Total courses found: {len(courses_list)}\n")

# Display first 10 courses
print("=== FIRST 10 COURSES ===\n")
for i, course in enumerate(courses_list[:10]):
    print(f"Course {i+1}:")
    print(f"  Name: {course.get('Course Name')}")
    print(f"  Category: {course.get('Category')}")
    print(f"  Duration: {course.get('Duration')}")
    print(f"  Price: {course.get('Price/Fees')}")
    print()

# Show summary of categories
categories = {}
for course in courses_list:
    cat = course.get('Category', 'Uncategorized')
    categories[cat] = categories.get(cat, 0) + 1

print("=== CATEGORIES SUMMARY ===\n")
for cat, count in sorted(categories.items()):
    print(f"{cat}: {count} courses")

# Save parsed data to JSON
with open('parsed_courses.json', 'w', encoding='utf-8') as f:
    json.dump(courses_list, f, indent=2, ensure_ascii=False)

print(f"\n✅ Saved {len(courses_list)} courses to parsed_courses.json")
